import csv
import io
import json

import openpyxl

FIELDS = [
    "Id_Tab",
    "Row",
    "Column",
    "Cod_Conto",
    "Cod_Dest2",
    "Cod_Dest3",
    "Cod_Dest4",
    "Cod_Dest5",
    "Cod_Categoria",
    "Formula",
    "Calculation_Logic",
    "DB_Storage_Sign",
    "EBA_Sign",
    "Coordinate",
]


def rgb(cell):
    return getattr(cell.fill.fgColor, "rgb", None)


def between(text, start, end):
    start_pos = text.find(start)
    end_pos = text.find(end, start_pos + len(start))

    if start_pos < 0 or end_pos < 0:
        return None

    return (
        text[start_pos + len(start): end_pos]
        .strip()
        .strip("'")
        .strip()
    )


def last(text, marker):
    start_pos = text.find(marker)

    if start_pos < 0:
        return None

    return (
        text[start_pos + len(marker):]
        .strip()
        .strip("'")
        .strip()
    )


def parse(value):
    if value is None:
        return None

    text = str(value).replace("\n", "")

    account = "Account= "
    dest2 = "Dest 2 = "
    dest3 = "Dest 3 = "
    dest4 = "Dest 4 = "
    dest5 = "Dest 5 = "
    category = "Category= "
    formula = "Formula= "
    storage_sign = "DB Storage Sign= "
    eba_sign = "EBA Sign= "
    calc_logic = "Calculation logic= "

    result = {
        key: None
        for key in FIELDS[3:13]
    }

    if (
        account in text
        and formula not in text
        and dest2 in text
    ):
        result.update(
            Cod_Conto=between(text, account, dest2),
            Cod_Dest2=between(text, dest2, dest3),
            Cod_Dest3=between(text, dest3, dest4),
            Cod_Dest4=between(text, dest4, dest5),
            Cod_Dest5=between(text, dest5, category),
            Cod_Categoria=between(text, category, storage_sign),
            DB_Storage_Sign=between(
                text,
                storage_sign,
                eba_sign
            ),
            EBA_Sign=(
                between(text, eba_sign, calc_logic)
                if calc_logic in text
                else last(text, eba_sign)
            ),
            Calculation_Logic=(
                last(text, calc_logic)
                if calc_logic in text
                else None
            ),
        )

    elif (
        account not in text
        and formula in text
        and calc_logic not in text
    ):
        result.update(
            Formula=between(
                text,
                formula,
                storage_sign
            ),
            DB_Storage_Sign=between(
                text,
                storage_sign,
                eba_sign
            ),
            EBA_Sign=last(text, eba_sign),
        )

    else:
        return None

    return (
        result
        if result["EBA_Sign"] is not None
        else None
    )


def process_workbook(path):
    workbook = openpyxl.load_workbook(
        path,
        data_only=False
    )

    sheet_names = workbook.sheetnames

    if not sheet_names:
        raise ValueError(
            "The Excel file does not contain valid worksheets."
        )

    reference_sheet = workbook[sheet_names[0]]

    color = rgb(reference_sheet["A1"])
    row_style = reference_sheet["A2"]._style
    column_style = reference_sheet["B1"]._style

    columns = {}
    rows = {}

    max_columns = {}
    max_rows = {}

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]

        current_column = 1
        current_row = 1

        for row in sheet.iter_rows():
            for cell in row:

                if cell.column == 1:
                    continue

                if (
                    cell.value is not None
                    and (
                        cell._style == column_style
                        or rgb(cell) == color
                    )
                ):
                    current_column += 1

                    columns[(sheet_name, current_column)] = (
                        "c" + str(cell.value).zfill(4)
                    )

                elif cell.value is None:
                    break

            break

        for column in sheet.iter_cols():
            for cell in column:

                if cell.row == 1:
                    continue

                if (
                    cell.value is not None
                    and (
                        cell._style == row_style
                        or rgb(cell) == color
                    )
                ):
                    current_row += 1

                    rows[(sheet_name, current_row)] = (
                        "r" + str(cell.value).zfill(4)
                    )

                elif cell.value is None:
                    break

            break

        max_columns[sheet_name] = current_column
        max_rows[sheet_name] = current_row

    records = []

    for sheet_name in sheet_names:

        if sheet_name in {
            "Test_Formula",
            "Macro"
        }:
            continue

        sheet = workbook[sheet_name]

        for row_range in sheet.iter_rows(
            min_row=2,
            max_row=max_rows[sheet_name],
            min_col=2,
            max_col=max_columns[sheet_name]
        ):
            for cell in row_range:

                if rgb(cell) == color:
                    continue

                parsed = parse(cell.value)

                if parsed:
                    records.append(
                        {
                            "Id_Tab": sheet_name,
                            "Row": rows.get(
                                (sheet_name, cell.row),
                                cell.row
                            ),
                            "Column": columns.get(
                                (sheet_name, cell.column),
                                cell.column
                            ),
                            **parsed,
                            "Coordinate": cell.coordinate,
                        }
                    )

    output = io.StringIO(newline="")

    writer = csv.DictWriter(
        output,
        fieldnames=FIELDS,
        lineterminator="\n"
    )

    writer.writeheader()
    writer.writerows(records)

    return json.dumps(
        {
            "records": records,
            "csv": output.getvalue(),
            "count": len(records),
        },
        ensure_ascii=False,
    )
