"""DPM 2.0 Taxonomy Extractor engine for Pyodide.

Ported from the desktop script dpm_2.0_json.py:
- the folder walk is replaced by the list of files provided by the browser;
- exclude_list and specific_list remain configurable parameters;
- the Excel workbook is written to the Pyodide virtual filesystem.

The extraction logic is preserved exactly as in the original script.
"""

import datetime
import json
import os
import re

import pandas as pd

DEFAULT_PATTERN = r"(?i)tab[\\/].*\.json$"

SHEET_ORDER = [
    "documentInfo_All",
    "factValue_All",
    "datapoint_dpm1_df_All",
    "datapoint_dpm1_agg_df_All",
    "datapoint_dpm2_df_All",
    "parameter_df_All",
    "parameter_allowedValue_df_All",
    "cell_allowedValue_df_All",
    "Mapping_dpm2_df_All",
]


def filter_paths(paths, pattern, exclude, specific):
    """Replicates trova_file_con_suffisso() using the browser file list."""
    regex = re.compile(pattern)
    selected = []

    for path in paths:
        path_lower = path.lower()

        matches_pattern = bool(regex.search(path))
        not_excluded = not any(e.lower() in path_lower for e in exclude if e.strip())
        is_specific = not specific or any(
            s.lower() in path_lower for s in specific if s.strip()
        )

        if matches_pattern and not_excluded and is_specific:
            selected.append(path)

    return selected


def json_tab_ext(file_path):
    """Extraction routine kept identical to the original desktop script."""
    with open(file_path) as f:
        data = json.load(f)

    documentInfo = pd.DataFrame(
        data["documentInfo"]["namespaces"].items(), columns=["namespace", "url"]
    )
    documentInfo["documentType"] = data["documentInfo"]["documentType"]

    tableTemplates_name = "tableTemplates"
    keys_table = list(data[tableTemplates_name].keys())

    datapoint_dpm1_df = pd.DataFrame()
    datapoint_dpm2_df = pd.DataFrame()
    parameter_df = pd.DataFrame()
    parameter_allowedValue_df = pd.DataFrame()
    cell_allowedValue = pd.DataFrame()
    cell_allowedValue_df = pd.DataFrame()
    decimals = pd.DataFrame()
    factValue = pd.DataFrame()

    for table in keys_table:
        keys_col = list(data[tableTemplates_name][table].keys())
        for col in keys_col:
            keys_col_element = list(data[tableTemplates_name][table][col].keys())
            for element in keys_col_element:
                if element == "factValue":
                    factValue = pd.DataFrame(
                        data[tableTemplates_name][table][col][element]["dimensions"].items(),
                        columns=["dimensions_ext", "dimensions"],
                    )
                    factValue["dimensions_code"] = factValue["dimensions"].str.replace(
                        "$", "", regex=False
                    )
                    factValue["propertiesFrom"] = data[tableTemplates_name][table][col][
                        element
                    ]["propertiesFrom"][0]
                elif element == "datapoint":
                    keys_datapoint = list(
                        data[tableTemplates_name][table][col][element][
                            "propertyGroups"
                        ].keys()
                    )
                    for datapoint in keys_datapoint:
                        keys_datapoint_elements = list(
                            data[tableTemplates_name][table][col][element][
                                "propertyGroups"
                            ][datapoint].keys()
                        )
                        for dp_element in keys_datapoint_elements:
                            if dp_element == "decimals":
                                decimals = pd.DataFrame(
                                    [
                                        data[tableTemplates_name][table][col][element][
                                            "propertyGroups"
                                        ][datapoint][dp_element]
                                    ],
                                    columns=["type_xbrl"],
                                )
                            elif dp_element == "dimensions":
                                datapoint_dimensions_dpm1 = pd.DataFrame(
                                    data[tableTemplates_name][table][col][element][
                                        "propertyGroups"
                                    ][datapoint][dp_element].items(),
                                    columns=["metric_dimension", "value"],
                                )
                            elif dp_element == "eba:documentation":
                                datapoint_dimensions_dpm2 = pd.DataFrame(
                                    [
                                        data[tableTemplates_name][table][col][element][
                                            "propertyGroups"
                                        ][datapoint][dp_element]
                                    ]
                                )
                                if "AllowedValue" in datapoint_dimensions_dpm2.columns:
                                    datapoint_dimensions_dpm2 = (
                                        datapoint_dimensions_dpm2.drop(
                                            columns=["AllowedValue"]
                                        )
                                    )
                                    cell_allowedValue = pd.DataFrame(
                                        data[tableTemplates_name][table][col][element][
                                            "propertyGroups"
                                        ][datapoint][dp_element]["AllowedValue"].items(),
                                        columns=["value", "code"],
                                    )
                                    cell_allowedValue["datapoint"] = datapoint
                                    decimals = pd.DataFrame(
                                        ["$textClosedList"], columns=["type_xbrl"]
                                    )
                                    cell_allowedValue_df = pd.concat(
                                        [cell_allowedValue_df, cell_allowedValue],
                                        axis=0,
                                        ignore_index=False,
                                    )
                        datapoint_dpm1_row = pd.concat(
                            [decimals, datapoint_dimensions_dpm1],
                            axis=1,
                            ignore_index=False,
                        )
                        datapoint_dpm1_row["datapoint"] = datapoint
                        datapoint_dpm1_df = pd.concat(
                            [datapoint_dpm1_df, datapoint_dpm1_row],
                            axis=0,
                            ignore_index=False,
                        )
                        datapoint_dpm2_row = pd.concat(
                            [decimals, datapoint_dimensions_dpm2],
                            axis=1,
                            ignore_index=False,
                        )
                        datapoint_dpm2_row["datapoint"] = datapoint
                        datapoint_dpm2_df = pd.concat(
                            [datapoint_dpm2_df, datapoint_dpm2_row],
                            axis=0,
                            ignore_index=False,
                        )
            for element in keys_col_element:
                if element in (factValue["dimensions_code"]).tolist():
                    keys_parameter = list(
                        data[tableTemplates_name][table][col][element].keys()
                    )
                    parameter_allowedValue = pd.DataFrame()
                    parameter_df_row_1 = pd.DataFrame()
                    parameter_df_row_2 = pd.DataFrame()
                    for parameter in keys_parameter:
                        if parameter == "eba:documentation":
                            parameter_df_row_1 = pd.DataFrame(
                                [data[tableTemplates_name][table][col][element][parameter]]
                            )
                            if "AllowedValue" in parameter_df_row_1.columns:
                                parameter_df_row_1 = parameter_df_row_1.drop(
                                    columns=["AllowedValue"]
                                )
                                parameter_allowedValue = pd.DataFrame(
                                    data[tableTemplates_name][table][col][element][
                                        parameter
                                    ]["AllowedValue"].items(),
                                    columns=["value", "code"],
                                )
                                parameter_allowedValue["parameter"] = element
                                parameter_allowedValue_df = pd.concat(
                                    [parameter_allowedValue_df, parameter_allowedValue],
                                    axis=0,
                                    ignore_index=False,
                                )
                        elif parameter == "tc:constraints":
                            parameter_df_row_2 = pd.DataFrame(
                                [data[tableTemplates_name][table][col][element][parameter]]
                            )
                            parameter_df_row_2 = parameter_df_row_2.rename(
                                columns={"type": "constraints_type"}
                            )
                    parameter_df_row = pd.concat(
                        [parameter_df_row_1, parameter_df_row_2],
                        axis=1,
                        ignore_index=False,
                    )
                    parameter_df_row["parameter"] = element
                    parameter_df = pd.concat(
                        [parameter_df, parameter_df_row], axis=0, ignore_index=False
                    )

        documentInfo["table"] = table
        factValue["table"] = table
        datapoint_dpm1_df["table"] = table
        datapoint_dpm2_df["table"] = table
        parameter_df["table"] = table
        parameter_allowedValue_df["table"] = table
        cell_allowedValue_df["table"] = table

        documentInfo["file_path"] = file_path
        factValue["file_path"] = file_path
        datapoint_dpm1_df["file_path"] = file_path
        datapoint_dpm2_df["file_path"] = file_path
        parameter_df["file_path"] = file_path
        parameter_allowedValue_df["file_path"] = file_path
        cell_allowedValue_df["file_path"] = file_path

        return (
            documentInfo,
            factValue,
            datapoint_dpm1_df,
            datapoint_dpm2_df,
            parameter_df,
            parameter_allowedValue_df,
            cell_allowedValue_df,
        )


def build_mapping(datapoint_dpm2_df_All, parameter_df_All):
    """Builds the Mapping_dpm2_df_All sheet."""
    if datapoint_dpm2_df_All.empty:
        return pd.DataFrame(
            columns=[
                "SheetVID",
                "FR_TABLE_U",
                "FR_TABLE_L",
                "FR_ROW",
                "FR_COLUMN",
                "datapoint",
                "parameter",
                "type",
                "headerCode",
            ]
        )

    required = ["SheetVID", "cellcode", "type", "datapoint", "table"]
    for column in required:
        if column not in datapoint_dpm2_df_All.columns:
            datapoint_dpm2_df_All[column] = None

    Mapping_dpm2_df_All = datapoint_dpm2_df_All[required].copy()

    tmp = (
        Mapping_dpm2_df_All["cellcode"]
        .astype("string")
        .str.strip("{}")
        .str.split(",", expand=True)
    )

    for index in range(3):
        if index not in tmp.columns:
            tmp[index] = None

    tmp = tmp.iloc[:, :3]
    tmp.columns = ["FR_TABLE_U", "FR_ROW", "FR_COLUMN"]
    tmp = tmp.apply(lambda column: column.str.strip())

    Mapping_dpm2_df_All[["FR_TABLE_U", "FR_ROW", "FR_COLUMN"]] = tmp

    Mapping_dpm2_df_All["FR_TABLE_L"] = (
        Mapping_dpm2_df_All["FR_TABLE_U"].str[0].str.lower()
        + Mapping_dpm2_df_All["FR_TABLE_U"].str[1:]
    )

    if not parameter_df_All.empty and {"table", "headerCode", "parameter"}.issubset(
        parameter_df_All.columns
    ):
        Parameter_tmp = parameter_df_All[["table", "headerCode", "parameter"]].copy()
        Parameter_tmp = (
            Parameter_tmp.sort_values(["table", "headerCode"])
            .groupby("table", as_index=False)
            .agg(
                {
                    "headerCode": lambda x: ",".join(x.astype(str)),
                    "parameter": lambda x: ",".join(x.astype(str)),
                }
            )
        )
        Mapping_dpm2_df_All = Mapping_dpm2_df_All.merge(
            Parameter_tmp, on="table", how="left"
        )
    else:
        Mapping_dpm2_df_All["headerCode"] = None
        Mapping_dpm2_df_All["parameter"] = None

    Mapping_dpm2_df_All.drop(
        columns=["cellcode", "table"], errors="ignore", inplace=True
    )

    return Mapping_dpm2_df_All[
        [
            "SheetVID",
            "FR_TABLE_U",
            "FR_TABLE_L",
            "FR_ROW",
            "FR_COLUMN",
            "datapoint",
            "parameter",
            "type",
            "headerCode",
        ]
    ]


def write_workbook(dfs, output_path):
    """Writes the workbook, preferring xlsxwriter and falling back to openpyxl."""
    try:
        import xlsxwriter  # noqa: F401

        with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
            workbook = writer.book
            header_format = workbook.add_format(
                {
                    "bold": True,
                    "bg_color": "#BBBBBB",
                    "font_name": "Calibri",
                    "font_size": 9,
                }
            )
            cell_format = workbook.add_format({"font_name": "Calibri", "font_size": 9})

            for sheet_name in SHEET_ORDER:
                df = dfs[sheet_name]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                worksheet.freeze_panes(1, 0)

                for index, column in enumerate(df.columns):
                    max_len = (
                        max(
                            df[column].astype(str).map(len).max() if not df.empty else 0,
                            len(str(column)),
                        )
                        + 2
                    )
                    max_len = min(max_len, 60)
                    worksheet.set_column(index, index, max_len, cell_format)
                    worksheet.write(0, index, str(column), header_format)

                n_rows, n_cols = df.shape
                if n_cols:
                    worksheet.autofilter(0, 0, n_rows, n_cols - 1)

        return "xlsxwriter"

    except ImportError:
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for sheet_name in SHEET_ORDER:
                df = dfs[sheet_name]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                worksheet.freeze_panes = "A2"

                fill = PatternFill("solid", fgColor="BBBBBB")
                font = Font(name="Calibri", size=9, bold=True)

                for index, column in enumerate(df.columns, start=1):
                    cell = worksheet.cell(row=1, column=index)
                    cell.fill = fill
                    cell.font = font

                    max_len = (
                        max(
                            df[column].astype(str).map(len).max() if not df.empty else 0,
                            len(str(column)),
                        )
                        + 2
                    )
                    worksheet.column_dimensions[get_column_letter(index)].width = min(
                        max_len, 60
                    )

                if df.shape[1]:
                    worksheet.auto_filter.ref = worksheet.dimensions

        return "openpyxl"


def process_taxonomy(
    file_paths_json,
    exclude_json,
    specific_json,
    output_path,
    pattern=DEFAULT_PATTERN,
):
    """Main entry point called from JavaScript."""
    start_time = datetime.datetime.now()

    all_paths = json.loads(file_paths_json)
    exclude = [item.strip() for item in json.loads(exclude_json) if item.strip()]
    specific = [item.strip() for item in json.loads(specific_json) if item.strip()]

    logs = ["=== DPM 2.0 Taxonomy Extractor started ==="]
    logs.append(f"JSON files received: {len(all_paths):,}")
    logs.append(f"Pattern: {pattern}")
    logs.append(f"Exclude list: {exclude if exclude else 'none'}")
    logs.append(f"Specific list: {specific if specific else 'none'}")

    selected = filter_paths(all_paths, pattern, exclude, specific)
    logs.append(f"Files matching the filters: {len(selected):,}")

    if not selected:
        raise ValueError(
            "No JSON file matched the pattern and the exclude/specific filters."
        )

    documentInfo_All = pd.DataFrame()
    factValue_All = pd.DataFrame()
    datapoint_dpm1_df_All = pd.DataFrame()
    datapoint_dpm2_df_All = pd.DataFrame()
    parameter_df_All = pd.DataFrame()
    parameter_allowedValue_df_All = pd.DataFrame()
    cell_allowedValue_df_All = pd.DataFrame()

    failed_files = []

    for path in selected:
        try:
            result = json_tab_ext(path)
            if result is None:
                failed_files.append(os.path.basename(path))
                continue

            (
                documentInfo,
                factValue,
                datapoint_dpm1_df,
                datapoint_dpm2_df,
                parameter_df,
                parameter_allowedValue_df,
                cell_allowedValue_df,
            ) = result

            documentInfo_All = pd.concat(
                [documentInfo_All, documentInfo], axis=0, ignore_index=False
            )
            factValue_All = pd.concat(
                [factValue_All, factValue], axis=0, ignore_index=False
            )
            datapoint_dpm1_df_All = pd.concat(
                [datapoint_dpm1_df_All, datapoint_dpm1_df], axis=0, ignore_index=False
            )
            datapoint_dpm2_df_All = pd.concat(
                [datapoint_dpm2_df_All, datapoint_dpm2_df], axis=0, ignore_index=False
            )
            parameter_df_All = pd.concat(
                [parameter_df_All, parameter_df], axis=0, ignore_index=False
            )
            parameter_allowedValue_df_All = pd.concat(
                [parameter_allowedValue_df_All, parameter_allowedValue_df],
                axis=0,
                ignore_index=False,
            )
            cell_allowedValue_df_All = pd.concat(
                [cell_allowedValue_df_All, cell_allowedValue_df],
                axis=0,
                ignore_index=False,
            )
        except Exception as error:
            failed_files.append(f"{os.path.basename(path)}: {error}")

    if "CellCode" in datapoint_dpm2_df_All.columns:
        if "cellcode" not in datapoint_dpm2_df_All.columns:
            datapoint_dpm2_df_All["cellcode"] = datapoint_dpm2_df_All["CellCode"]
        else:
            datapoint_dpm2_df_All["cellcode"] = datapoint_dpm2_df_All[
                "CellCode"
            ].fillna(datapoint_dpm2_df_All["cellcode"])

    datapoint_dpm2_df_All = datapoint_dpm2_df_All.drop(
        columns=["CellCode"], errors="ignore"
    )

    if not datapoint_dpm1_df_All.empty:
        datapoint_dpm1_agg_df_All = datapoint_dpm1_df_All.groupby(
            ["datapoint", "table", "file_path"], as_index=False, dropna=False
        ).agg({"metric_dimension": "/".join, "value": "/".join})
    else:
        datapoint_dpm1_agg_df_All = pd.DataFrame()

    Mapping_dpm2_df_All = build_mapping(datapoint_dpm2_df_All, parameter_df_All)

    dfs = {
        "documentInfo_All": documentInfo_All,
        "factValue_All": factValue_All,
        "datapoint_dpm1_df_All": datapoint_dpm1_df_All,
        "datapoint_dpm1_agg_df_All": datapoint_dpm1_agg_df_All,
        "datapoint_dpm2_df_All": datapoint_dpm2_df_All,
        "parameter_df_All": parameter_df_All,
        "parameter_allowedValue_df_All": parameter_allowedValue_df_All,
        "cell_allowedValue_df_All": cell_allowedValue_df_All,
        "Mapping_dpm2_df_All": Mapping_dpm2_df_All,
    }

    engine = write_workbook(dfs, output_path)

    sheet_summary = [
        {"sheet": name, "rows": int(len(dfs[name])), "columns": int(dfs[name].shape[1])}
        for name in SHEET_ORDER
    ]

    for entry in sheet_summary:
        logs.append(
            f"Sheet {entry['sheet']}: {entry['rows']:,} rows, {entry['columns']} columns"
        )

    if failed_files:
        logs.append(f"Files skipped: {len(failed_files)}")
        for item in failed_files[:20]:
            logs.append(f"Skipped: {item}")

    duration = datetime.datetime.now() - start_time
    logs.append(f"Excel engine: {engine}")
    logs.append(f"Total duration: {duration}")
    logs.append("=== Extraction completed ===")

    preview = Mapping_dpm2_df_All.head(200).copy()
    preview = preview.where(pd.notna(preview), None)

    return json.dumps(
        {
            "output_path": output_path,
            "processed_files": len(selected) - len(failed_files),
            "matched_files": len(selected),
            "received_files": len(all_paths),
            "skipped_files": failed_files,
            "mapping_rows": int(len(Mapping_dpm2_df_All)),
            "sheets": sheet_summary,
            "preview": preview.to_dict(orient="records"),
            "columns": list(Mapping_dpm2_df_All.columns),
            "logs": logs,
        },
        ensure_ascii=False,
        default=str,
    )
